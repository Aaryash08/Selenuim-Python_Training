import pandas as pd
import openpyxl
df = pd.read_excel('sales_data.xlsx', sheet_name='2025')

df['Total'] = df['Quantity'] * df['Price']

df.to_excel('sales_summary1.xlsx', index=False)

print("Pandas: File saved successfully!")


wb = openpyxl.load_workbook('sales_data.xlsx')
sheet = wb['2025']

sheet.cell(row=1, column=4).value = "Total"

for row in range(2, sheet.max_row + 1):
    quantity = sheet.cell(row=row, column=2).value
    price = sheet.cell(row=row, column=3).value
    if quantity is not None and price is not None:
        sheet.cell(row=row, column=4).value = quantity * price

wb.save('sales_summary2.xlsx')
print("OpenPyXL: File saved successfully!")